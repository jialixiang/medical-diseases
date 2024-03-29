#!/usr/bin/python
# -*- coding: utf-8 -*-

import csv
from openpyxl import Workbook, load_workbook

from config import chronic, epidemic, infection

def check_fuzzy_match(string, array):
    for entry in array:
        if entry in string or string in entry:
            return True

    return False


if __name__ == '__main__':

    files = [
        u"内科学第八版",
        u"妇产科学第八版",
        u"皮肤性病学第六版",
        u"儿科学第八版",
        # PDF converted
        u"外科学八年制第二版",
        u"临床医学概论",
        u"其他",
    ]

    wb = Workbook()
    wb = load_workbook(filename='files/疾病信息表 - 20170410 - v1 - for Jiali.xlsx')
    ws = wb['Sheet1']

    results = {}
    for filename in files:
        with open('files/output/%s.csv' % filename) as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                key = row.pop('\xe7\x96\xbe\xe7\x97\x85\xe5\x90\x8d\xe7\xa7\xb0')
                if not results.has_key(key):
                    results[key] = row

    unknown = set()

    matched_count = 0
    for index, row in enumerate(ws.iter_rows()):
        target_name = row[0].value
        if not target_name or target_name == u'疾病名称':
            continue

        gender = ''
        age = ''

        # 检查慢性病、流行病、传染病
        is_chronic = u"是" if (target_name in chronic or check_fuzzy_match(target_name, chronic)) else u"否"
        ws.cell(row=index+1, column=3, value=is_chronic)

        is_epidemic = u"是" if (target_name in epidemic or check_fuzzy_match(target_name, epidemic)) else u"否"
        ws.cell(row=index+1, column=6, value=is_epidemic)

        is_infection = u"是" if (target_name in infection or check_fuzzy_match(target_name, infection)) else u"否"
        ws.cell(row=index+1, column=5, value=is_infection)

        target_name = target_name.encode('utf-8')

        if results.has_key(target_name):
            # 完全匹配
            matched_count += 1
            print "%s / %s" % (target_name, target_name)
            gender = results[target_name]['\xe6\x80\xa7\xe5\x88\xab']
            age = results[target_name]['\xe5\xb9\xb4\xe9\xbe\x84\xe6\xae\xb5']
        else:
            # 部分匹配
            for name in results:
                if name in target_name or target_name in name:
                    print "%s / %s" % (name, target_name)
                    gender = results[name]['\xe6\x80\xa7\xe5\x88\xab']
                    age = results[name]['\xe5\xb9\xb4\xe9\xbe\x84\xe6\xae\xb5']
                    matched_count += 1
                    break

        if not (gender or age):
            unknown.add(target_name)

        try:
            ws.cell(row=index+1, column=9, value=gender)
            ws.cell(row=index+1, column=10, value=age)
        except Exception, e:
            print "ERROR", e
            pass

    wb.save('output.xlsx')

    print
    print "Matched in total:", matched_count
    print
    print "Unknown in total:", len(unknown)
    for name in unknown:
        print name
